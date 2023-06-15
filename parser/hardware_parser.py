from openpyxl import load_workbook

JSON_OUT = {}

wb = load_workbook(filename="Аудиторный-фонд-_для-оснащения_.xlsx")


def main():
    ws = wb['Оснащены оборудованием']
    for row in ws.iter_rows(min_row=2, max_row=58, min_col=2, max_col=23, values_only=True):
        arr = row[2].replace(" ", "").split("+")
        JSON_OUT[row[0]] = {
            "equipment": arr,
            "comment": None
        }
    print(JSON_OUT)


if __name__ == "__main__":
    main()
