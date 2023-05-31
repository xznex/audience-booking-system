import json
import os
from typing import List

from openpyxl import load_workbook

PATH = os.path.dirname(__file__) + r"\src"

GRADUATE_SCHOOL = 1

JSON_OUT = {}


def get_files(path: str) -> List:
    files = []

    for filename in os.listdir(path):

        f = os.path.join(path, filename)

        if os.path.isfile(f):
            files.append(f)

    return files


def get_merged_cells(cur_sheet, start_with_b) -> List:
    merged_cells = []

    for merged_cell in cur_sheet.merged_cells.ranges:
        if merged_cell.min_col == start_with_b + 2 and merged_cell.max_col == start_with_b + 2:
            merged_cells.append(merged_cell)

    return merged_cells


def is_merged_cell(cur_sheet, col, row) -> bool:
    is_merged = False

    for merged_cell in cur_sheet.merged_cells.ranges:
        if merged_cell.min_col == col and merged_cell.max_col == col \
                and (merged_cell.min_row == row or merged_cell.max_row == row):
            is_merged = True

    return is_merged


def parsing(wb, i, week_parity_flag=False):
    sheet = wb.worksheets[i]
    start_with_b = 0

    if sheet.title == "ИГУиП 1 курс":
        start_with_b = 1

    MAX_COLS = sheet.max_column

    JSON_OUT[sheet.title] = {}

    merged_cells = get_merged_cells(sheet, start_with_b)

    if merged_cells:
        for day_coord in merged_cells:
            _, bottom, _, top = day_coord.bounds
            day_of_week = sheet[bottom][1 + start_with_b].value
            JSON_OUT[sheet.title][day_of_week] = {}
            delay = top - bottom + 1
            for i in range(delay):
                if i % 2 == 0:
                    period = sheet[bottom + i][2 + start_with_b].value
                    JSON_OUT[sheet.title][day_of_week][period] = {}

                    merged_rows = []
                    for k in range(MAX_COLS - 4):
                        is_merged = is_merged_cell(sheet, k + start_with_b + 5, bottom + i)

                        if is_merged:
                            merged_rows.append(k + start_with_b + 5)

                    for j in range(2):
                        week_parity = sheet[bottom + i + j][3 + start_with_b].value

                        if week_parity_flag:
                            if j == 0:
                                week_parity = "ЧЁТ."
                            else:
                                week_parity = "НЕЧЁТ."

                        JSON_OUT[sheet.title][day_of_week][period][week_parity] = {}
                        subjects = []

                        for k in range(MAX_COLS - 4 - start_with_b):
                            # same_val_above = False
                            if j == 1 and (k + start_with_b + 5 in merged_rows):
                                subject = sheet[bottom + i][k + start_with_b + 4].value
                            else:
                                subject = sheet[bottom + i + j][k + start_with_b + 4].value

                            if subject is None:
                                continue

                            subject = subject.replace("\n", " ")
                            subjects.append(subject)

                        JSON_OUT[sheet.title][day_of_week][period][week_parity] = subjects
                        # print(excel_path, sheet.title, day_of_week, period, week_parity, subjects)


for excel_path in get_files(PATH):
    try:
        wb = load_workbook(filename=excel_path)
    except PermissionError as e:
        print("Необходимо закрыть все файлы Excel для корректной работы.")

    quantity_sheets = len(wb.sheetnames)

    if quantity_sheets == GRADUATE_SCHOOL:
        ws = wb.active

        if ws['D6'].value == "НЕД.":
            parsing(wb, 0)
        else:
            parsing(wb, 0, True)
    else:
        for i in range(quantity_sheets):
            parsing(wb, i)

with open("data_file.json", "w") as write_file:
    json.dump(JSON_OUT, write_file, ensure_ascii=False)
